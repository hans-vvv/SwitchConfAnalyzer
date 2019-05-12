import re
from collections import defaultdict
import tkinter as tk
from tkinter import filedialog
from openpyxl.utils import get_column_letter
from openpyxl import Workbook


class ReSearcher:

    """
    Helper class to enable evaluation
    and regex formatting in a single line
    """

    match = None

    def __call__(self, pattern, string):
        self.match = re.search(pattern, string)
        return self.match

    def __getattr__(self, name):
        return getattr(self.match, name)


def xlref(row, column, zero_indexed=True):
    """
    openpyxl helper
    """
    if zero_indexed:
        row += 1
        column += 1
    return get_column_letter(column) + str(row)



def get_value(key, item):

    """
    key + value = item
    function return value for given key and item
    """

    if key.strip() == item.strip():
        return key
    else:
        item = item.lstrip()
        result = re.search('^('+key+')(.*)', item)
        value = format(result.group(2)).lstrip()
        return value



def get_key(item, key_length):

    """
    key + value = item
    number of words of key = key_length
    function returns key
    """

    word = item.strip().split()
    if key_length == 0: # fix
        return item
    elif len(word) == key_length:
        return item
    else:
        return ' '.join(word[0:key_length])

def splitrange(raw_range):

    """
    ex. splitrange('105-107') will return ['105','106','107']
    """

    result = []

    if  re.search(r'^(\d+)\-(\d+)$', raw_range):
        match = re.search(r'^(\d+)\-(\d+)$', raw_range)
        first = int(format(match.group(1)))
        last = int(format(match.group(2)))
        for i in range(first, last+1):
            result.append(str(i))
        return result



def get_switch_info(file_path):

    """
    This function stores interface and Vlan specific parts of switch
    configurations in a (nested) dictionary.
    For the following keys, lists are used to store values:
    - switchport trunk allowed vlan (add)
    - standby (HSRP)
    - ip helper-address
    Known Caveats:
    - No support for secundairy ip addresses
    - Only Vlan name is read under VLAN configuration
    - If SVI of VLAN exists but VLAN doesn't then VLAN index appears to be
      present in result.
    """

    # Helper function
    def store_port_items(line, vlanindex, portindex):

        """
        This helper function stores interface items.
        The following methods are used in the given order to determine
        which part of an interface item is considered to be a key and which
        part a value.

        1. First portkey_exceptions list is examined. If an interface item
           contains the words found in this list then key = item in the list
           and value = remaining words of the interface item. If an interface
           item is found then the other methods are not considered.
        2. Portkeys dict is examined. If interface item contains an item
           found in a list of the dict then corresponding key (i.e. 1 or 2)
           is used to split the item. The key of the item is equal to the
           number of words of the dict key, the rest of the item = value.
           Example: If line = channel-group 2 mode active, then
           key = "channel-group"  and value = "2 mode active". If an interface
           item is found then the last method is not considered.
        3. Default method. Last word of line = value
           and all preceding words = key.
        """

        portkey_exceptions = ['ip vrf forwarding']

        portkeys = {1: ['hold-queue', 'standby', 'channel-group', 'description'],
                    2: ['switchport port-security', 'ip', 'spanning-tree',
                        'speed auto', 'srr-queue bandwidth']}

        line = line.lstrip()
        found_item = False

        # 'Method 1'
        for key in portkey_exceptions:
            if key in line:
                if 'Vlan' in portindex:
                    vlaninfo[vlanindex][key] = get_value(key, line)
                    found_item = True
                else:
                    portinfo[portindex][key] = get_value(key, line)
                    found_item = True

        # 'Method 2'
        for key_length in portkeys:
            if found_item:
                continue
            for item in portkeys[key_length]:
                if item not in line:
                    continue
                key = get_key(line, key_length)
                if 'standby' in line:
                    if 'Vlan' in portindex:
                        standby.append(get_value(key, line))
                        vlaninfo[vlanindex]['standby'] = ','.join(standby)
                        found_item = True
                    else:
                        standby.append(get_value(key, line))
                        portinfo[portindex]['standby'] = ','.join(standby)
                        found_item = True
                elif 'ip helper-address' in line:
                    if 'Vlan' in portindex:
                        ip_helper.append(get_value(key, line))
                        helper = ','.join(ip_helper)
                        vlaninfo[vlanindex]['ip helper-address'] = helper
                        found_item = True
                    else:
                        ip_helper.append(get_value(key, line))
                        helper = ','.join(ip_helper)
                        portinfo[portindex]['ip helper-address'] = helper
                        found_item = True
                elif 'Vlan' in portindex:
                    vlaninfo[vlanindex][key] = get_value(key, line)
                    found_item = True
                else:
                    portinfo[portindex][key] = get_value(key, line)
                    found_item = True

        # 'Method 3 or default method'
        if not found_item:
            key = get_key(line, len(word)-1)
            if 'switchport trunk allowed vlan' in line:
                for raw_vlans in get_value(key, line).split(','):
                    if '-' in raw_vlans:
                        for vlan_id in splitrange(raw_vlans):
                            vlan_allow_list.append(vlan_id)
                    else:
                        vlan_allow_list.append(raw_vlans)
                portinfo[portindex]['vlan_allow_list'] = ','.join(vlan_allow_list)
            elif 'Vlan' in portindex:
                vlaninfo[vlanindex][key] = get_value(key, line)
            else:
                portinfo[portindex][key] = get_value(key, line)



    # Start main part of function
    switchinfo = defaultdict(dict) # Dict containing all info

    portinfo = defaultdict(dict)
    vlaninfo = defaultdict(dict)
    scanfile = False

    match = ReSearcher()

    with open(file_path, 'r') as lines:

        for line in lines:

            line = line.rstrip()
            word = line.split()

            if match(r'^interface (Vlan(\d+))', line):
                scanfile = True
                portindex = format(match.group(1))
                vlanindex = format(match.group(2))
                vlaninfo[vlanindex]['vlanindex'] = vlanindex
                standby = []
                ip_helper = []

            elif match(r'^interface (.*)', line):
                scanfile = True
                portindex = format(match.group(1))
                portinfo[portindex]['portindex'] = portindex
                vlan_allow_list = []
                standby = []
                ip_helper = []

            elif match(r'^vlan (\d+)\-(\d+)$', line):
                scanfile = True
                start_vlan = int(match.group(1))
                stop_vlan = int(match.group(2))
                for vlan in range(start_vlan, stop_vlan+1):
                    vlaninfo[str(vlan)]['vlanindex'] = str(vlan)

            elif match(r'^vlan (\d+)$', line):
                scanfile = True
                vlanindex = format(match.group(1))
                vlaninfo[vlanindex]['vlanindex'] = vlanindex

            elif match(r'^ name (.*)', line) and scanfile:
                vlaninfo[vlanindex]['name'] = format(match.group(1))

            elif match(r'^ no (.*)', line) and scanfile:
                key = format(match.group(1))
                value = format(match.group(0))
                if 'Vlan' in portindex:
                    vlaninfo[vlanindex][key] = value
                else:
                    portinfo[portindex][key] = value

            elif match(r'^hostname (.*)', line):
                hostname = format(match.group(1))

            elif match(r'!$', line) and scanfile:
                scanfile = False

            # interface items are stored with helper function
            elif match('^ .*', line) and scanfile:
                store_port_items(line, vlanindex, portindex)

                
        switchinfo['portinfo'] = portinfo
        switchinfo['vlaninfo'] = vlaninfo
        switchinfo['generalinfo']['hostname'] = hostname

    return switchinfo


def calc_vlan_use(switchinfo):

    """
    This function returns Vlan usage statistics of the switch.
    Helpfull if switch is true (stub) access switch.
    """

    vlans = []
    access_vlans = []
    for vlan, vlanitems in switchinfo['vlaninfo'].items():
        vlans.append(vlan)

    for port, portitems in switchinfo['portinfo'].items():
        if portitems.get('switchport access vlan') is None:
            continue
        vlan = portitems['switchport access vlan']
        access_vlans.append(vlan)
            
    access_vlans = sorted(set(access_vlans), key=int)     
    vlans = sorted(set(vlans), key=int)
    diff = sorted(set(vlans).symmetric_difference(set(access_vlans)), key=int)

    if '1' in diff:
        diff.remove('1')

    result = []

    fmt1 = 'The following VLANs are present on the switch: '
    fmt2 = "The following VLANs are present on all access ports: "
    fmt3 = 'The following VLANs are candidate to be removed: '

    result.append(fmt1 + ','.join(vlans))
    result.append(fmt2 + ','.join(access_vlans))
    result.append(fmt3 + ','.join(diff))

    return result



def info_to_xls(switchinfo):

    """
    Function print switchinfo object to excel file. The primary
    keys of switchinfo object are printed in seperated tabs.
    """

    # Calculate list of port- and vlan keys
    vlankeys = []
    portkeys = []

    for vlan, vlanitems in switchinfo['vlaninfo'].items():
        for key in vlanitems.keys():
            vlankeys.append(key)

    for port, portitems in switchinfo['portinfo'].items():
        for key in portitems.keys():
            portkeys.append(key)

    vlankeys = sorted(set(vlankeys))
    portkeys = sorted(set(portkeys))
    vlankeys.remove('name')
    portkeys.remove('description')
    portkeys.remove('portindex')
    vlankeys.remove('vlanindex')
    vlankeys.insert(0, 'name')
    portkeys.insert(0, 'description')
    
    wb = Workbook()
    ws = wb.create_sheet("Vlaninfo", 0)
    ws = wb.create_sheet("Portinfo", 0)
    ws = wb.create_sheet("Vlan_report", 0)

    ws = wb['Vlaninfo']

    count_vlan_row = 0
    ws[xlref(0, 0)] = 'vlan'

    for count, vlankey in enumerate(vlankeys):
        ws[xlref(0, count+1)] = vlankey

    for vlan, vlanitems in switchinfo['vlaninfo'].items():
        ws[xlref(count_vlan_row+1, 0)] = vlan

        for count_col, vlankey in enumerate(vlankeys):
            value = vlanitems.get(vlankey, '')
            ws[xlref(count_vlan_row+1, count_col+1)] = value
        count_vlan_row += 1


    ws = wb['Portinfo']

    count_port_row = 0
    ws[xlref(0, 0)] = 'interface'

    for count, portkey in enumerate(portkeys):
        ws[xlref(0, count+1)] = portkey

    for port, portitems in switchinfo['portinfo'].items():
        ws[xlref(count_port_row+1, 0)] = port

        for count_col, portkey in enumerate(portkeys):
            value = portitems.get(portkey, '')
            ws[xlref(count_port_row+1, count_col+1)] = value
        count_port_row += 1


    ws = wb['Vlan_report']

    for count, line in enumerate(calc_vlan_use(switchinfo)):
        ws[xlref(count, 0)] = line

    wb.save(switchinfo['generalinfo']['hostname'] + '.xlsx')


def add_interface_properties(switchinfo):

    """
    Interface properties of switch are inserted in switchinfo object based on
    configuration of all ports.
    """

    #Add interface properties (access, ip, trunk, unused) to switchinfo object.
    for port, portitems in switchinfo['portinfo'].items():

        if 'switchport access vlan' in portitems.keys():
            switchinfo['portinfo'][port]['mode'] = 'access'

        elif portitems.get('switchport mode') == 'trunk':
            switchinfo['portinfo'][port]['mode'] = 'trunk'

        elif portitems.get('switchport') == 'no switchport':
            switchinfo['portinfo'][port]['mode'] = 'ip'

        # management interface
        elif portitems.get('ip address') == 'no ip address':
            switchinfo['portinfo'][port]['mode'] = 'unused'

        # management interface
        elif portitems.get('ip address') is not None:
            switchinfo['portinfo'][port]['mode'] = 'ip'

        else:
            switchinfo['portinfo'][port]['mode'] = 'unused'


    for vlan, vlanitems in switchinfo['vlaninfo'].items():

        if vlanitems.get('ip address') == 'no ip address':
            switchinfo['vlaninfo'][vlan]['mode'] = 'ip'

        elif vlanitems.get('ip address') is not None:
            switchinfo['vlaninfo'][vlan]['mode'] = 'ip'

    return switchinfo
    


def read_config_template(config_template_file):

    """
    Config template file is read. The file consists of general, hierarchical
    and interface configuration parts. For each type of interface
    (IP, access, trunk) a template can be made; also items can be excluded
    for comparison with switch config. Listed items in hierarchical config
    parts must be separate by ! sign. Also after the last item a ! must
    be present.

    """

    #Read config template and store in object.
    template_dict = defaultdict(list)
    with open(config_template_file, 'r') as lines:

        hierarc_scan = False # State var to store hierarchical config parts
        hierarc_temp = [] # store hierarc config item

        for line in lines:
            line = line.rstrip()

            if not line.strip(): # ignore empty lines
                continue

            if line == '# Access interface items':
                scan_item = 'acc_intf'

            elif line == '# Access interface items ignore':
                scan_item = 'acc_intf_ign'

            elif line == '# Trunk interface items':
                scan_item = 'trk_intf'

            elif line == '# Trunk interface items ignore':
                scan_item = 'trk_intf_ign'

            elif line == '# IP interface items':
                scan_item = 'ip_intf'

            elif line == '# IP interface items ignore':
                scan_item = 'ip_intf_ign'

            elif line == '# Global config items':
                scan_item = 'glob'

            elif line == '# Global items beginning ignore':
                scan_item = 'glob_begin_ign'

            elif line == '# Global items subset ignore':
                scan_item = 'glob_subset_ign'

            elif line == '# Global hierarchical config items':
                hierarc_scan = True

            if '#' in line:
                continue

            if hierarc_scan:
                if line == '!':
                    # append list in list ...
                    template_dict['global_hierarc'].append(hierarc_temp)
                    hierarc_temp = []
                else:
                    hierarc_temp.append(line)
            else:
                template_dict[scan_item].append(line)


    return template_dict


def gen_intf_comparison(switchinfo, template_dict):

    """
    Differences between interface specific switch configuration and template
    are printed. A configuration is produced to shutdown all unused ports.
    """

    # Open file to report complaincy results
    switchname = switchinfo['generalinfo']['hostname']
    comparison_file = switchname  + '-comparison-result.txt'

    with open(comparison_file, 'w') as res:

        # Create config to shutdown interfaces if it has not been shutdown.
        print('################ Shutdown unused interfaces:', file=res)
        for port, portitems in switchinfo['portinfo'].items():

            if (portitems.get('mode') == 'unused'
                    and portitems.get('shutdown') != 'shutdown'):

                print('interface ' + port, file=res)
                print(' shutdown', file=res)
                print('!', file=res)

        # Print access, trunk and IP complaincy reports.
        comp_intf_info = [('access', 'acc_intf_ign', 'acc_intf'),
                          ('trunk', 'trk_intf_ign', 'trk_intf'),
                          ('ip', 'ip_intf_ign', 'ip_intf')]

        for intf_type, comparison_intf_ign, comparison_intf in comp_intf_info:
            print('!', file=res)
            fmt = '################ Configure {} ports according template:'
            print(fmt.format(intf_type), file=res)
            for port, portitems in switchinfo['portinfo'].items():
                found = [] # Relevant configured items under interfaces

                if portitems.get('mode') != intf_type:
                    continue

                # Do not analyze member intf's of Port-ch's
                if 'channel-group' in portitems.keys():
                    continue

                print('interface ' + port, file=res)

                for k, v in portitems.items():

                    if k == 'switchport access vlan':
                        continue
                    if k == 'mode':
                        continue
                    if k == 'portindex':
                        continue
                    if k in template_dict[comparison_intf_ign]:
                        continue

                    v = v.lstrip()
                    v_list = v.split()
                    if k == v:
                        found.append(k)
                    elif k == ' '.join(v_list[1:]):
                        found.append(v)
                    else:
                        found.append(k + ' ' + v)

                for item in template_dict[comparison_intf]:
                    if item not in found:
                        print(item, file=res)

                for item in found:
                    if (item not in template_dict[comparison_intf]
                            and item not in template_dict[comparison_intf_ign]):
                        print('no ' + item, file=res)
                print('!', file=res)



        for vlan, vlanitems in switchinfo['vlaninfo'].items():

            if vlanitems.get('mode') != 'ip':
                continue

            print('interface Vlan' + vlan, file=res)
            found = [] # Relevant configured items under interfaces
            for k, v in vlanitems.items():

                if k == 'name':
                    continue
                if k == 'mode':
                    continue
                if k == 'vlanindex':
                    continue
                if k in template_dict['ip_intf_ign']:
                    continue

                v = v.lstrip()
                value = v.split()
                if k == v:
                    found.append(k)
                elif k == ' '.join(value[1:]):
                    found.append(v)
                else:
                    found.append(k + ' ' + v)

            for item in template_dict['ip_intf']:
                if item not in found:
                    print(item, file=res)
            for item in found:
                if (item not in template_dict['ip_intf_ign']
                        and item not in template_dict['ip_intf']):
                    print('no ' + item, file=res)
            print('!', file=res)


    return comparison_file


def gen_hier_config_part_names(file_path):

    """
    This function makes a list out of all first lines of all hierarchical config
    parts. This list is used to filter out parts of the switch configuration
    in the next function so that only hierarchical config parts in config
    template are compared to hierarchical config parts in the switch
    configuration.
    """

    with open(file_path, 'r') as lines:

        hier_config_part_names = []
        temp = ''
        hier_var = False
        for line in lines:

            if not line.strip(): # skip empty lines
                continue
            line = line.rstrip()

            if line == line.lstrip():
                fmt = '^(interface|vlan|banner|ip access-list)(.*)'
                if re.search(fmt, line): # hard code exceptions:
                    hier_config_part_names.append(line)
                else:
                    temp = line

            if line != line.lstrip() and not hier_var:
                hier_config_part_names.append(temp)
                hier_var = True
                temp = ''

            if hier_var and line == '!':
                hier_var = False

    return hier_config_part_names




def gen_audit_config(template_dict, hier_config_part_names, file_path):

    """
    This function filters out items from switch config. The result is compared
    to config template in last function.
    """


    with open(file_path, 'r') as lines:
        skipline = False # If true, config part is ignored
        audit_config = [] # Includes global config items
        temp_hierarc = [] # Store hierarchical config part
        audit_hierarc_config = [] # List of hierarchical config parts
        hier_var = False # If true hierarchical config is stored

        global_hierarc_templ_items = []
        for item in template_dict['global_hierarc']:
            # First line of all lists
            global_hierarc_templ_items.append(item[0])

        for line in lines:

            if not line.strip():
                continue
            line = line.rstrip()
            words = line.split()

            if hier_var:
                if line != line.lstrip():
                    temp_hierarc.append(line)
                elif line == line.lstrip():
                    audit_hierarc_config.append(temp_hierarc)
                    hier_var = False
                elif line == '!':
                    audit_hierarc_config.append(temp_hierarc)
                    hier_var = False

            for item in global_hierarc_templ_items:
                if line == item:
                    hier_var = True
                    temp_hierarc = []
                    temp_hierarc.append(line)

            for item in hier_config_part_names:
                if item == line:
                    skipline = True

            if skipline or hier_var:
                if line == '!':
                    skipline = False
            else:
                filterline = False
                for item in template_dict['glob_begin_ign']:
                    result = re.search('^('+item+')*', line)
                    if result.group(0):
                        filterline = True
                for item in template_dict['glob_subset_ign']:
                    items = item.split()
                    if set(items).issubset(words):
                        filterline = True
                if not filterline:
                    audit_config.append(line)


##        for config in audit_config:
##            print(config)
##
##        print(audit_hierarc_config)


    return audit_config, audit_hierarc_config


def gen_general_differences(comparison_file, template_dict,
                            audit_config, audit_hierarc_config):

    """
    This function presents differences between switch configuration and
    configuration template. The result presents suggestions to remediate
    the differences.
    """

    with open(comparison_file, 'a') as res:

        non_compl_items = list(set(audit_config) - set(template_dict['glob']))
        print('################ Non compliant General items:', file=res)
        for item in sorted(non_compl_items):
            words = item.split()
            if words[0].lstrip() != 'no':
                print('no ' + item, file=res)
            else:
                print(' '.join(words[1:]), file=res)
        print('!', file=res)
        print('################ Missing general template items:', file=res)
        for item in template_dict['glob']:
            if item not in audit_config:
                print(item, file=res)

        print('################ Non compliant hierarchical items:', file=res)
        global_hierarc_templ_items = []
        for item in template_dict['global_hierarc']:
            # First line of all lists
            global_hierarc_templ_items.append(item[0])

        global_hier_config_items = []
        for item in audit_hierarc_config:
            # First line of all lists
            global_hier_config_items.append(item[0])

        print('!', file=res)
        diff = set(global_hierarc_templ_items) - set(global_hier_config_items)
        if diff:
            fmt = ('Difference(s) found in hierarchical config sections. '
                   'Present in template, not in config:')
            print(fmt, file=res)
            for item in diff:
                for item1 in template_dict['global_hierarc']:
                    if item1[0] == item:
                        for item in item1:
                            print(item, file=res)
                        print('!', file=res)

        # Remove list not found in config
        for index, item in enumerate(template_dict['global_hierarc']):
            if item[0] in diff:
                template_dict['global_hierarc'].pop(index)

        template_dict['global_hierarc'] = sorted(template_dict['global_hierarc'],
                                                 key=lambda x: x[0])

        audit_hierarc_config = sorted(audit_hierarc_config, key=lambda x: x[0])

        #print(template_dict['global_hierarc'])

        for hierconf_part, hierconf_configpart \
                in zip(template_dict['global_hierarc'], audit_hierarc_config):
            if hierconf_part != hierconf_configpart:
                print('Difference found in {}.'.format(
                    hierconf_part[0]), file=res)
                print('!', file=res)
                print('Template configuration is:', file=res)
                for item in hierconf_part:
                    print(item, file=res)
                print('!', file=res)
                print('Switch configuration is:', file=res)
                for item in hierconf_configpart:
                    print(item, file=res)
            print('!', file=res)
            print('!', file=res)
            print('!', file=res)


def main():

    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename() # switch configuration

    # Retrieve interface and vlan info from configuration file and store in
    # switchinfo object.
    switchinfo = get_switch_info(file_path)

    # Print switchinfo object in excel file.
    info_to_xls(switchinfo)

    # Add interface properties (access, trunk, IP, unused)
    add_interface_properties(switchinfo)

    # Read config template and store in object.
    config_template_file = 'config_template.txt'
    template_dict = read_config_template(config_template_file)

    # Report interface specific differences with user defined template.
    comparison_file = gen_intf_comparison(switchinfo, template_dict)

    # Generate list with first line of all hierarchical config parts.
    hier_config_part_names = gen_hier_config_part_names(file_path)

    # Filter config items from switch config. Result will be compared
    # with template items in next function.
    audit_config, audit_hierarc_config = gen_audit_config(
        template_dict, hier_config_part_names, file_path)

    # Calculate and present differences between switch config and template.
    gen_general_differences(
        comparison_file, template_dict, audit_config, audit_hierarc_config)


main()



































